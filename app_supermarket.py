import streamlit as st
import pandas as pd
import sqlite3
import os
import json
import math
import io
import altair as alt
import base64
import openpyxl
import hashlib
import plotly.express as px
import streamlit_highcharts as hg
#from reportlab.pdfgen import canvas
from datetime import datetime , timedelta
#from reportlab.lib.pagesizes import letter
import streamlit.components.v1 as components
from fpdf import FPDF
from fpdf.enums import XPos, YPos
import time
import streamlit_antd_components as sac


#if os.path.exists('dblite/supermarket.db'):
#    os.remove('dblite/supermarket.db')

# --- 1. CORE ENGINE ---
def get_connection():
    return sqlite3.connect('dblite/supermarket.db', check_same_thread=False)

def hash_password(password):
    # Mengubah password menjadi hash SHA-256
    return hashlib.sha256(str.encode(password)).hexdigest()

# Contoh cara memasukkan user baru ke database (jalankan sekali saja)
def add_user(username, password, role):
    conn = get_connection()
    cursor = conn.cursor()
    hashed_p = hash_password(password)
    cursor.execute("INSERT INTO users (username, password, role) VALUES (?, ?, ?)", 
                   (username, hashed_p, role))
    conn.commit()
    conn.close()



# --- 3. UI MODULES ---



def hex_to_rgba(hex_str, opacity=0.3):
    hex_str = hex_str.lstrip('#')
    r, g, b = tuple(int(hex_str[i:i+2], 16) for i in (0, 2, 4))
    return f'rgba({r}, {g}, {b}, {opacity})'


def apply_theme():
    # Hitung warna teks adaptif

    if 'theme' not in st.session_state: return
    t = st.session_state.get('theme', {
        'bg_color': '#0E1117',
        'top_bar': '#95A5A6',
        'primary_color': '#00FFA3',
        'font_family': 'Segoe UI',
        'font_size': '14px'
    })
    sidebar_text = get_contrast_color(t['bg_color'])
    body_text = get_contrast_color(t['body_color'])
    btn_text_hover = get_contrast_color(t['bg_color'])

    
    st.markdown(f"""
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Orbitron:wght@400;700&family=Roboto:wght@300;400;700&display=swap');
        


        /* 1. TOP BAR (HEADER) */
        header[data-testid="stHeader"] {{
            background-color: {t['top_bar']} !important; /* Disamakan dengan body atau beri warna lain */            
            /*border-bottom: 1px solid rgba(255, 255, 255, 0.1);*/
        }}

        /* 1. WARNA HALAMAN UTAMA (BODY) */
        .stApp {{
            background-color: {t['body_color']} !important;
            color: white;
            font-family: '{t['font_family']}', sans-serif;
        }}

        /* 2. WARNA SIDEBAR (DIBEDAKAN) */
        [data-testid="stSidebar"] {{
            background-color: {t['bg_color']} !important;
            border-right: 1px solid rgba(255, 255, 255, 0.1);
        }}

        /* Futuristic Card Effect */
        div[data-testid="stMetricValue"] {{
            color: var(--primary);
            text-shadow: 0 0 5px var(--primary);
        }}

        /* --- 1. KOTAK FORM LOGIN (NEON BORDER) --- */
        [data-testid="stForm"] {{
            border: 2px solid {t['bg_color']} !important;
            border-radius: 5px !important;
            padding: 30px !important;
            background-color: rgba(255, 255, 255, 0.03) !important; /* Efek Glassmorphism */
            box-shadow: 0 0 10px rgba({hex_to_rgba(t['bg_color'])}, 0.3) !important;
            backdrop-filter: blur(5px); /* Efek Blur di belakang kotak */
        }}

        /* --- 2. INPUT FIELD DALAM FORM --- */
        .stTextInput input {{
            background-color: rgba(0, 0, 0, 0.2) !important;
            color: white !important;
            border: 1px solid rgba({hex_to_rgba(t['bg_color'])}, 0.3) !important;
            border-radius: 8px !important;
        }}
        
        /* TARGET SEMUA TOMBOL */
        div.stButton > button, 
        div.stDownloadButton > button, 
        button[kind="secondaryFormSubmit"] {{
            /* BACKGROUND HARUS GELAP/TRANSPARAN AGAR TEKS TERANG TERLIHAT */
            background-color: rgba({hex_to_rgba(t['bg_color'])}, 0.3) !important; 
            color: {t['primary_color']} !important;
            
            border: 2px solid {t['primary_color']} !important;
            border-radius: 8px !important;
            padding: 10px !important;
            font-weight: 700 !important;
            font-family: 'Orbitron', sans-serif !important;
            text-transform: uppercase;
            
            /* Efek bayangan agar teks lebih 'tebal' */
            text-shadow: 0 0 5px rgba(0,0,0,0.5);
            transition: all 0.3s ease;
        }}

        /* SAAT KURSOR DI ATAS TOMBOL */
        div.stButton > button:hover, 
        div.stDownloadButton > button:hover, 
        button[kind="secondaryFormSubmit"]:hover {{
            background-color: {t['primary_color']} !important;
            color: {btn_text_hover} !important; /* Otomatis Hitam jika primary cerah */
            box-shadow: 0 0 20px {t['primary_color']} !important;
        }}

        /* MENGHILANGKAN BACKGROUND PUTIH BAWAAN ANDROID/MOBILE */
        div.stButton > button:active, div.stButton > button:focus {{
            background-color: #0a101f !important;
            color: {t['primary_color']} !important;
        }}

        
        


        /* 3. INPUT FIELD (Agar tetap terbaca) */
        .stTextInput input, .stSelectbox select {{
            background-color: rgba(255,255,255,0.1) !important;
            color: {body_text} !important;
        }}

        /* 4. FIX ICON PANAH (Agar tidak muncul teks) */
        [data-testid="stIcon"], .st-emotion-cache-6qob1r, [data-testid="collapsedControl"] {{
            font-family: serif !important;
            color: {sidebar_text} !important;
        }}

             
        </style>
    """, unsafe_allow_html=True)



  
# --- 3. PROMO LOGIC ---
def calculate_promos(cart):
    total_before = sum(item['Subtotal'] for item in cart)
    discount_total = 0
    promo_details = []

    for item in cart:
        # Contoh Promo: Buy 2 Get 1 Free untuk Susu Steril
        if item['Product'] == 'Susu Steril' and item['Qty'] >= 3:
            free_units = item['Qty'] // 3
            discount_val = free_units * item['Price']
            discount_total += discount_val
            promo_details.append(f"Promo B2G1 Susu: -{discount_val:,.0f}")

    # Contoh Promo: Diskon Global 5% jika belanja > 200rb
    if total_before >= 200000:
        global_disc = total_before * 0.05
        discount_total += global_disc
        promo_details.append(f"Diskon Member 5%: -{global_disc:,.0f}")

    return discount_total, promo_details

# --- 4. UI MODULES ---
# --- 3. HELPERS ---
def create_receipt_text(t_id, date, cashier, items, total):
    receipt = f"======================\n       FUTURE MART 2026         \n   NEO-RETAIL POS SYSTEM        \n======================\nNo. Inv : {t_id}\nTanggal : {date}\nKasir   : {cashier}\n----------------------\n"
    for i in items:
        receipt += f"{i['Product'][:20]:<20}\n  {i['Qty']} x {i['Price']:,.0f}      {i['Subtotal']:>10,.0f}\n"
    receipt += "--------------------------------\n"
    receipt += f"TOTAL           Rp {total:>10,.0f}\n======================\n  TERIMA KASIH TELAH BERBELANJA \n      POWERED BY KENYED      \n"
    return receipt
    
# --- 3. UI MODULES ---




#Auto Backup Cloud
def show_database_tools():
    st.markdown("<h2 style='color:var(--primary); font-family:Orbitron;'>🛡️ DATA PROTECTION</h2>", unsafe_allow_html=True)
    st.markdown('<div class="glass-card">', unsafe_allow_html=True)
    
    st.subheader("Manual Cloud Backup")
    st.write("Unduh salinan database Anda secara berkala untuk disimpan di Google Drive atau Hardisk Eksternal.")
    
    # Membaca file database SQLite
    with open("dblite/supermarket.db", "rb") as f:
        db_bytes = f.read()
    
    st.download_button(
        label="📥 DOWNLOAD DATABASE (.db)",
        data=db_bytes,
        file_name=f"FutureMart_Backup_{datetime.now().strftime('%Y%m%d_%H%M')}.db",
        mime="application/octet-stream",
        width='stretch'
    )
    
    st.markdown("---")
    st.subheader("Data Cleanup")
    st.warning("Hanya gunakan fitur ini saat pergantian tahun fiskal atau jika database terlalu berat.")
    
    if st.button("🗑️ RESET TRANSACTION HISTORY", help="Menghapus semua riwayat transaksi tapi menyimpan data produk"):
        conn = get_connection()
        c = conn.cursor()
        c.execute("DELETE FROM trans_detail")
        c.execute("DELETE FROM trans_master")

        conn.commit()
        st.success("Riwayat transaksi telah dibersihkan!")
    
    st.markdown('</div>', unsafe_allow_html=True)




def graph_packedbubble():
    if 'theme' not in st.session_state: return
    t = st.session_state.get('theme', {
        'bg_color': '#0E1117',
        'top_bar': '#95A5A6',
        'primary_color': '#00FFA3',
        'font_family': 'Segoe UI',
        'font_size': '14px'
    })
    sidebar_text = get_contrast_color(t['bg_color'])
    body_text = get_contrast_color(t['body_color'])
    btn_text_hover = get_contrast_color(t['bg_color'])
    warna_bg = get_contrast_color(t['top_bar'])

    st.markdown(f"<h5 style='color:{t['primary_color']}; font-family:Orbitron;'>Product, make by Quantity</h5>", unsafe_allow_html=True)

    conn = get_connection()    
    #zx = pd.read_sql("select product_name, category,sum(qty) as jumlah from trans_detail group by product_name,category", conn)
    query = "select category,product_name,sum(qty) as jumlah from trans_detail group by category, product_name"
    zx = pd.read_sql(query, conn)
    
    #result= query.fetch_veh_graph()
    df=pd.DataFrame(zx, columns=['product_name','category','jumlah'])
    # 2. Transform DataFrame into Highcharts series format
    chart_data = []
    for category in df['category'].unique():
        subset = df[df['category'] == category]
        data_points = subset[['category','product_name', 'jumlah']].rename(columns={'category': 'category','product_name': 'name', 'jumlah': 'value'}).to_dict('records')
        chart_data.append({'name': category, 'data': data_points})
        #data_points = subset[['MAKE', 'JUMLAH']].rename(columns={'MAKE': 'make', 'JUMLAH': 'jumlah'}).to_dict('records')
        #chart_data = [{'product_name': point['product_name'],'name': point['name'], 'value': point['value']} for point in data_points]
        
    # 3. Define Chart Configuration
    chartDef = { 'chart': { 'height': '60%', 'height': '60%',
                'type': 'packedbubble',
                'backgroundColor': t['body_color']},    
                'title': { 
                    'text': 'Product and Category', # <-- Judul Anda di sini
                    'style': {
                        'color': 'black',                 # Opsional: Sesuaikan warna teks judul
                        'fontWeight': 'bold'
                    }
                },                            
    'plotOptions': { 'packedbubble': { 'dataLabels': { 'enabled': True,
                                                        #'filter': { 'operator': '>',
                                                        #            'property': 'y'
                                                        #            'value': 10},
                                                        'format': '{point.name}',
                                                        'style': { 'color': 'black',
                                                                    'fontWeight': 'normal',
                                                                    'textOutline': 'none',
                                                                    'backgroundColor':t['top_bar']}},
                                        'layoutAlgorithm': { 'dragBetweenSeries': True,
                                                            'gravitationalConstant': 0.08,
                                                            'parentNodeLimit': True,
                                                            'seriesInteraction': False,
                                                            'splitSeries': True},
                                        'maxSize': '100%',
                                        'minSize': '20%',
                                        'zMax': 1000,
                                        'zMin': 0}},
    'series': chart_data
    }            
    data=hg.streamlit_highcharts(chartDef,300)
    st.markdown("##")
    return data  




def get_transaction_trends():
    conn = get_connection()
    
    # Mengambil total penjualan per tanggal
    query=("""
        SELECT date as tanggal, SUM(total) as total 
        FROM trans_master 
        GROUP BY tanggal 
        ORDER BY tanggal DESC LIMIT 7
    """)
  
    df = pd.read_sql(query, conn)
    conn.close()
    return df

def show_transaction_charts():
    df_trend = get_transaction_trends()
    
    
    if not df_trend.empty:

        #st.subheader("📈 Tren Penjualan (7 Hari Terakhir)")
        t = st.session_state.theme
        st.markdown(f"<h5 style='color:{t['primary_color']}; font-family:Orbitron;'>📈 Penjualan (7 Hari Terakhir)</h5>", unsafe_allow_html=True)

        # Pengaturan Warna Tema
        neon_color = "#00f2ff" # Cyan Neon
        grid_color = "#31333f" # Warna garis tipis agar tidak mengganggu
        
        chart = alt.Chart(df_trend).mark_area(
            line={'color': neon_color},
            color=alt.Gradient(
                gradient='linear',
                stops=[alt.GradientStop(color=neon_color, offset=0),
                       alt.GradientStop(color='transparent', offset=1)],
                x1=1, x2=1, y1=1, y2=0
            ),
            interpolate='monotone' # Membuat garis melengkung halus
        ).encode(
            x=alt.X('tanggal:T', title='Tanggal', axis=alt.Axis(grid=False, labelAngle=-45, labelColor='black', titleColor='black')),
            y=alt.Y('total:Q', title='Total Penjualan (Rp)', axis=alt.Axis(grid=True, gridColor=grid_color, labelColor='black', titleColor='black')),
            tooltip=['tanggal', 'total']
        ).properties(
            height=300,
            background='transparent' # Menyatu dengan tema Streamlit
        ).configure_view(
            strokeWidth=0 # Menghapus border kotak
        ).interactive()

        st.altair_chart(chart, width='stretch')
    else:
        st.info("Belum ada data transaksi untuk ditampilkan.")

def show_top_products_chart():
    conn = get_connection()
    query = "SELECT product_name, SUM(qty) as jumlah FROM trans_detail GROUP BY product_name ORDER BY jumlah DESC LIMIT 5"
    zx = pd.read_sql(query, conn)
    conn.close()

    t = st.session_state.theme    
    st.markdown(f"<h5 style='color:{t['primary_color']}; font-family:Orbitron;'>🏆 Top 5 Produk Terlaris</h5>", unsafe_allow_html=True)

    if not zx.empty:
        #st.subheader("🏆 Top 5 Produk Terlaris")
                
        bar_chart = alt.Chart(zx).mark_bar(
            cornerRadiusTopLeft=10,
            cornerRadiusTopRight=10
        ).encode(
            x=alt.X("product_name:N", sort="-y", title="Nama Produk", 
                    axis=alt.Axis(labelColor='black', titleColor='black', labelAngle=-45)),
            y=alt.Y("jumlah:Q", title="Jumlah Terjual", 
                    axis=alt.Axis(labelColor='black', titleColor='black')),
            # Menambahkan warna berbeda untuk setiap produk
            color=alt.Color("product_name:N", legend=None, scale=alt.Scale(scheme='tableau20')),
            tooltip=['product_name', 'jumlah']
        ).properties(
            height=300,
            background='transparent'
        ).configure_view(
            strokeOpacity=0
        )

        st.altair_chart(bar_chart, use_container_width=True, theme=None)

def graph_bar():
    
    if 'theme' not in st.session_state: return
    t = st.session_state.get('theme', {
        'bg_color': '#0E1117',
        'top_bar': '#95A5A6',
        'primary_color': '#00FFA3',
        'font_family': 'Segoe UI',
        'font_size': '14px'
    })
    sidebar_text = get_contrast_color(t['bg_color'])
    body_text = get_contrast_color(t['body_color'])
    btn_text_hover = get_contrast_color(t['bg_color'])
    warna_bg = get_contrast_color(t['top_bar'])
    asli_top_bar = t['top_bar']

    conn = get_connection()    

    query = "select category,sum(qty) as jumlah from trans_detail group by category"
    zx = pd.read_sql(query, conn)

    t = st.session_state.theme    
    st.markdown(f"<h5 style='color:{t['primary_color']}; font-family:Orbitron;'>Product by Quantity</h5>", unsafe_allow_html=True)

    #st.subheader("Product by Quantity")
    source = pd.DataFrame({
    "Quantity ($)": zx["jumlah"],
    "Product Name": zx["category"]
    })

    # Tentukan warna background sesuai tema (misal: transparan atau biru gelap)
    bg_color = "transparent" # atau gunakan kode hex seperti "#0e1117"
    text_color = warna_bg # "#FFFFFF"   # Warna teks agar kontras

    bar_chart = alt.Chart(source).mark_bar(
        # Hapus 'color=asli_top_bar' dari sini
        cornerRadiusEnd=5     # Membuat ujung bar melengkung (elegan)
    ).encode(
        x=alt.X("sum(Quantity ($)):Q", title="Total Quantity", axis=alt.Axis(labelColor=text_color, titleColor=text_color)),
        y=alt.Y("Product Name:N", sort="-x", title="Product", axis=alt.Axis(labelColor=text_color, titleColor=text_color)),
        
        # --- TAMBAHKAN KEMBALI COLOR DI SINI UNTUK WARNA OTOMATIS/GRADIENT ---
        color=alt.Color("Product Name:N", legend=None) # Menggunakan nama produk sebagai kategori warna
        # Anda juga bisa menambahkan .scale(scheme='rainbow') untuk skema warna spesifik
    ).properties(
        width='container',    # Menyesuaikan lebar
        height=300,
        background=bg_color   # MENGATUR BACKGROUND DI SINI
    ).configure_view(
        strokeOpacity=0       # Menghilangkan garis tepi kotak grafik
    ).configure_axis(
        grid=False            # Menghilangkan garis kotak-kotak di belakang (lebih bersih)
    )

    st.altair_chart(bar_chart, width='stretch')




def show_dashboard():
    t = st.session_state.theme    
    st.markdown(f"<h4 style='color:{t['primary_color']}; font-family:Orbitron;'>CORE DASHBOARD</h4>", unsafe_allow_html=True)
    
    t = st.session_state.theme
    conn = get_connection()    

    coll1,coll2=st.columns(2)

    with coll1:
        
        show_top_products_chart()
        graph_packedbubble() 

      
        


    with coll2:    
        #st.subheader("Product and Make by Quantity")
        show_transaction_charts()         
        graph_bar()
 

        

        

    

def barcode_scanner():
    # Komponen HTML/JS untuk akses kamera
    scanner_code = """
    <div id="interactive" class="viewport" style="width: 100%; height: 300px; position: relative;">
        <video id="video" style="width: 100%; height: 100%; object-fit: cover;"></video>
        <canvas class="drawingBuffer" style="display: none;"></canvas>
    </div>
    <script src="https://cdn.jsdelivr.net/npm/@ericblade/quagga2/dist/quagga.min.js"></script>
    <script>
        Quagga.init({
            inputStream: {
                name: "Live",
                type: "LiveStream",
                target: document.querySelector('#interactive'),
                constraints: { facingMode: "environment" } // Gunakan kamera belakang
            },
            decoder: {
                readers: ["ean_reader", "code_128_reader", "ean_8_reader"]
            }
        }, function(err) {
            if (err) { console.log(err); return; }
            Quagga.start();
        });

        Quagga.onDetected(function(result) {
            var code = result.codeResult.code;
            // Kirim kode barcode kembali ke Python Streamlit
            window.parent.postMessage({
                type: 'streamlit:setComponentValue',
                value: code
            }, '*');
            Quagga.stop();
        });
    </script>
    """
    # Menangkap hasil scan dari JavaScript
    st.markdown("### 📷 Scan Barcode")
    scan_result = components.html(scanner_code, height=350)
    return scan_result

# Gunakan di menu POS
def show_pos_with_scanner():
    st.markdown("<h6 style='color:var(--primary); font-family:Orbitron;'>🛒 Kamera</h6>", unsafe_allow_html=True)
    
    # Toggle untuk membuka kamera
    if st.checkbox("Buka Scanner Kamera"):
        barcode = barcode_scanner()
        if barcode:
            st.success(f"Barcode Terdeteksi: {barcode}")
            # Cari produk di database berdasarkan barcode
            # query = "SELECT * FROM products WHERE barcode = ?"
    
    # Sisanya adalah logika transaksi seperti sebelumnya

#CETAK PRINTER TERMAL 58mm atau 80mm
def print_receipt_thermal(receipt_text):
    # Menggunakan iframe tersembunyi untuk mencetak agar tidak merusak tampilan utama
    receipt_html = f"""
    <html>
    <body onload="window.print();">
        <div id="receipt" style="font-family: 'Courier New', Courier, monospace; width: 200px; font-size: 12px;">
            <pre style="white-space: pre-wrap; margin: 0;">{receipt_text}</pre>
        </div>
        <script>
            // Event ini dipicu setelah dialog print ditutup (di-print atau di-cancel)
            window.onafterprint = function() {{
                // Mengirim pesan ke Streamlit bahwa proses selesai (opsional)
                console.log("Print selesai/ditutup");
            }};
        </script>
    </body>
    </html>
    """
    # Gunakan height sedikit lebih besar agar tidak ada scrollbar di dalam iframe cetak
    components.html(receipt_html, height=80)

def show_pos():
    t = st.session_state.theme    
    conn = get_connection()
    
    # Init session state cart jika belum ada
    if 'cart' not in st.session_state:
        st.session_state.cart = []

    # --- HEADER & INFO TRANSAKSI ---
    st.markdown('<div class="glass-card">', unsafe_allow_html=True)
    header_col1, header_col2, header_col3 = st.columns([1.5, 1.5, 1])
    with header_col1:
        t_id_preview = f"INV-{datetime.now().strftime('%y%m%d%H%M%S')}"
        st.text_input("Invoice ID", t_id_preview, disabled=True)
    with header_col2:
        m_search = st.text_input("No. Member (Opsional)", placeholder="0812345678")
    with header_col3:
        st.text_input("Cashier", st.session_state.username, disabled=True)
    
    # Validasi Member
    member_data = None
    if m_search:
        res = pd.read_sql("SELECT * FROM members WHERE phone=?", conn, params=[m_search])
        if not res.empty:
            member_data = res.iloc[0]
            st.markdown(f'''<div style="background: rgba(0, 212, 255, 0.1); padding: 10px; border-radius: 5px; border-left: 5px solid #00d4ff; margin-bottom:10px;">
                <span style="color: #00d4ff; font-weight: bold;">⭐ MEMBER:</span> {member_data["name"]} | <b>Poin: {member_data["points"]}</b>
            </div>''', unsafe_allow_html=True)
        else:
            st.error("⚠️ Member tidak ditemukan")
    st.markdown('</div>', unsafe_allow_html=True)

        # --- INPUT PRODUK ---
    df_p = pd.read_sql("SELECT * FROM products WHERE stock > 0", conn)
    st.markdown('<div class="glass-card">', unsafe_allow_html=True)
    
    # Mengatur rasio kolom agar Qty dan Tombol lebih pas (kecil)
    col_prod, col_price, col_qty, col_btn = st.columns([2.5, 1, 0.8, 1])
    
    with col_prod:
        sel_p = st.selectbox("Cari Produk", df_p['name'], index=None, placeholder="Pilih item...")
    
    if sel_p:
        p_data = df_p[df_p['name'] == sel_p].iloc[0]
        with col_price:
            st.metric("Harga", f"{p_data['price']:,.0f}")
        
        with col_qty:
            # Komponen Qty
            qty = st.number_input("Qty", min_value=1, max_value=int(p_data['stock']), value=1)
            
        with col_btn:
            # Trik: Memberikan spasi vertikal agar tombol turun sejajar dengan input Qty
            st.markdown('<div style="margin-top: 28px;"></div>', unsafe_allow_html=True)
            if st.button("➕ TAMBAH", use_container_width=True, type="primary"):
                st.session_state.cart.append({
                    'Product': sel_p, 
                    'Category': p_data['category'], 
                    'Qty': qty, 
                    'Price': p_data['price'], 
                    'Subtotal': qty * p_data['price']
                })
                st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)

    # --- KERANJANG BELANJA & MANAJEMEN ITEM ---
    if st.session_state.cart:
        #st.markdown('<div class="glass-card">', unsafe_allow_html=True)
        #st.subheader("🛒 Detail Pesanan")
        
        # Header Tabel Custom
        h_col1, h_col2, h_col3, h_col4, h_col5 = st.columns([2, 0.5, 1, 1, 0.5])
        h_col1.caption("PRODUK")
        h_col2.caption("QTY")
        h_col3.caption("HARGA")
        h_col4.caption("TOTAL")
        h_col5.caption("AKSI")
        
        subtotal = 0
        for idx, item in enumerate(st.session_state.cart):
            b_col1, b_col2, b_col3, b_col4, b_col5 = st.columns([2, 0.5, 1, 1, 0.5])
            b_col1.write(f"**{item['Product']}**")
            b_col2.write(str(item['Qty']))
            b_col3.write(f"{item['Price']:,.0f}")
            b_col4.write(f"{item['Subtotal']:,.0f}")
            if b_col5.button("🗑️", key=f"del_{idx}"):
                st.session_state.cart.pop(idx)
                st.rerun()
            subtotal += item['Subtotal']

        st.divider()
        
        # --- PEMBAYARAN & DISKON ---
        pay_col1, pay_col2 = st.columns([2, 1])
        final_discount = 0
        
        with pay_col1:
            if member_data and member_data['points'] > 0:
                use_points = st.checkbox(f"Gunakan Poin sebagai Diskon? (Maks: {member_data['points']})")
                if use_points:
                    final_discount = member_data['points']
                    st.info(f"Potongan Poin: -Rp {final_discount:,.0f}")
            
            if st.button("🚫 KOSONGKAN KERANJANG", type="secondary"):
                st.session_state.cart = []
                st.rerun()

        grand_total = subtotal - final_discount
        with pay_col2:
            st.markdown(f"<h3 style='text-align:right; margin:0;'>TOTAL AKHIR</h3>", unsafe_allow_html=True)
            st.markdown(f"<h1 style='text-align:right; color:#00d4ff; margin:0;'>Rp {grand_total:,.0f}</h1>", unsafe_allow_html=True)
        
        st.write("##")
        if st.button("🚀 PROSES PEMBAYARAN SEKARANG", use_container_width=True, type="primary"):
            c = conn.cursor()
            t_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Simpan Transaksi
            c.execute("INSERT INTO trans_master VALUES (?,?,?,?,?,?)", 
                      (t_id_preview, t_date, grand_total, st.session_state.username, final_discount, m_search))

            new_points_earned = int(grand_total * 0.01)
            for i in st.session_state.cart:
                c.execute("UPDATE products SET stock = stock - ? WHERE name = ?", (i['Qty'], i['Product']))
                c.execute("INSERT INTO trans_detail VALUES (?,?,?,?,?,?,?)", (t_id_preview, i['Product'], i['Category'], i['Qty'], i['Price'], i['Subtotal'], 0))
            
            if member_data is not None:
                point_change = new_points_earned - (final_discount if use_points else 0)
                c.execute("UPDATE members SET points = points + ? WHERE phone = ?", (point_change, m_search))
            
            conn.commit()
            st.session_state.last_receipt = create_receipt_text(t_id_preview, t_date, st.session_state.username, st.session_state.cart, subtotal)
            st.session_state.cart = []
            st.success(f"Berhasil! Member mendapatkan +{new_points_earned} poin.")
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    # --- STRUK PREVIEW ---
    if 'last_receipt' in st.session_state and st.session_state.last_receipt:
        with st.expander("📄 PREVIEW STRUK", expanded=True):
            st.code(st.session_state.last_receipt)
            c1, c2, c3 = st.columns(3)
            with c1:
                if st.button("🖨️ CETAK FISIK", use_container_width=True):
                    print_receipt_thermal(st.session_state.last_receipt)
            with c2:
                st.download_button("📥 DOWNLOAD TXT", data=st.session_state.last_receipt, file_name=f"Receipt_{datetime.now().strftime('%H%M%S')}.txt", use_container_width=True)
            with c3:
                if st.button("TUTUP", use_container_width=True): 
                    st.session_state.last_receipt = None
                    st.rerun()

def get_all_users():
    conn = get_connection()
    df = pd.read_sql("SELECT username, role FROM users", conn)
    conn.close()
    return df

    
def add_new_user(username, password, role):
    conn = get_connection()
    cursor = conn.cursor()
    try:
        hashed_p = hash_password(password)
        cursor.execute("INSERT INTO users (username, password, role) VALUES (?, ?, ?)", 
                       (username, hashed_p, role))
        conn.commit()
        return True, "User berhasil ditambahkan!"
    except Exception as e:
        return False, f"Error: {e}"
    finally:
        conn.close()

def delete_user(username):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM users WHERE username = ?", (username,))
    conn.commit()
    conn.close()

# Fungsi untuk mencatat aktivitas
def add_log(username, action):
    conn = get_connection()
    cursor = conn.cursor()
    # Menggunakan waktu lokal Indonesia
    cursor.execute("INSERT INTO user_logs (username, action, timestamp) VALUES (?, ?, datetime('now', 'localtime'))", 
                   (username, action))
    conn.commit()
    conn.close()

_="""
def show_user_mgmt():
    st.markdown("<h2 style='color:var(--primary); font-family:Orbitron;'>👥 USER MANAGEMENT</h2>", unsafe_allow_html=True)
    conn = get_connection()
    df_u = pd.read_sql("SELECT username, role FROM users", conn)
    
    st.markdown('<div class="glass-card">', unsafe_allow_html=True)
    st.write("Existing Users")
    st.dataframe(df_u, width='stretch')
    
    with st.expander("Add New User"):
        new_u = st.text_input("New Username")
        new_p = st.text_input("New Password", type="password")
        new_r = st.selectbox("Role", ["Admin", "Cashier"])
        if st.button("Create User"):               
            add_user(new_u, new_p, new_r)
            st.success("User Created!")
            st.success("User Created!")
            st.rerun()            
           
    st.markdown('</div>', unsafe_allow_html=True)
"""

def show_activity_logs():
    t = st.session_state.theme
    st.markdown(f"<h2 style='color:{t['primary_color']}; font-family:Orbitron;'>🖥️ SYSTEM ACTIVITY LOG</h2>", unsafe_allow_html=True)
    


    conn = get_connection()
    #query = conn.cursor()
    # Ambil 50 aktivitas terbaru
    query =("SELECT username as USER, action as AKTIVITAS, timestamp as WAKTU FROM user_logs ORDER BY id DESC LIMIT 50")
    df_logs = pd.read_sql(query, conn)
    conn.close()

    if not df_logs.empty:
        # Menampilkan log dengan styling neon
        st.dataframe(df_logs, width='stretch')
        
        if st.button("🗑️ BERSIHKAN LOG LAMA"):
            # Opsi untuk menghapus log jika sudah terlalu penuh
            conn = get_connection()
            conn.execute("DELETE FROM activity_logs")
            conn.commit()
            conn.close()
            st.rerun()
    else:
        st.info("Belum ada aktivitas yang tercatat.")


def show_user_mgmt():
    t = st.session_state.theme
    st.markdown(f"<h2 style='color:{t['primary_color']}; font-family:Orbitron;'>👥 USER MANAGEMENT</h2>", unsafe_allow_html=True)
    
    # Bagian 1: List User Terdaftar
    st.subheader("Daftar Akun")
    users_df = get_all_users()
    st.dataframe(users_df, width='stretch')
    
    # Bagian 2: Tambah User Baru
    with st.expander("➕ TAMBAH USER BARU"):
        with st.form("form_add_user", clear_on_submit=True):
            new_u = st.text_input("Username Baru")
            new_p = st.text_input("Password", type="password")
            new_r = st.selectbox("Role", ["Admin", "Cashier"])
            
            if st.form_submit_button("SIMPAN USER"):
                if new_u and new_p:
                    success, msg = add_new_user(new_u, new_p, new_r)
                    if success:
                        st.success(msg)
                        st.rerun()
                    else:
                        st.error(msg)
                else:
                    st.warning("Isi semua kolom!")

    # Bagian 3: Hapus User
    with st.expander("🗑️ HAPUS USER"):
        user_to_del = st.selectbox("Pilih User untuk Dihapus", users_df['username'].tolist())
        if st.button("KONFIRMASI HAPUS", type="primary"):
            if user_to_del == "admin": # Proteksi agar admin utama tidak terhapus
                st.error("Admin utama tidak boleh dihapus!")
            else:
                delete_user(user_to_del)
                st.success(f"User {user_to_del} telah dihapus.")
                st.rerun()

def save_theme(theme_dict):
    with open("theme_config.json", "w") as f:
        json.dump(theme_dict, f)

def load_theme():
    if os.path.exists("theme_config.json"):
        with open("theme_config.json", "r") as f:
            return json.load(f)
    return None # Jika belum ada file, gunakan default

def get_contrast_color(hex_color):
    """Mengembalikan 'white' untuk background gelap dan 'black' untuk background terang"""
    hex_color = hex_color.lstrip('#')
    r, g, b = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
    
    # Rumus Standar Luminance (YIQ)
    luminance = (r * 0.299 + g * 0.587 + b * 0.114) / 255
    return "white" if luminance < 0.5 else "#1e1e1e"

def generate_invoice_pdf(master_row, detail_df):
    pdf = FPDF()
    pdf.add_page()
    
    # Header Struk
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(0, 10, "FUTUREMART 2026", ln=True, align='C')
    pdf.set_font("Arial", '', 10)
    pdf.cell(0, 5, "Sistem Kasir Futuristik Terintegrasi", ln=True, align='C')
    pdf.ln(10)
    
    # Informasi Transaksi
    pdf.set_font("Arial", 'B', 10)
    pdf.cell(0, 8, f"ID INVOICE : {master_row['id']}", ln=True)
    pdf.set_font("Arial", '', 10)
    pdf.cell(0, 8, f"Tanggal    : {master_row['date']}", ln=True)
    pdf.cell(0, 8, f"Kasir      : {master_row['cashier']}", ln=True)
    pdf.ln(5)
    
    # Tabel Header
    pdf.set_fill_color(200, 200, 200)
    pdf.cell(80, 8, "Nama Produk", 1, 0, 'C', True)
    pdf.cell(20, 8, "Qty", 1, 0, 'C', True)
    pdf.cell(40, 8, "Harga", 1, 0, 'C', True)
    pdf.cell(45, 8, "Subtotal", 1, 1, 'C', True)
    
    # Tabel Detail
    for _, row in detail_df.iterrows():
        pdf.cell(80, 8, str(row['product_name']), 1)
        pdf.cell(20, 8, str(row['qty']), 1, 0, 'C')
        pdf.cell(40, 8, f"Rp {row['price']:,.0f}", 1, 0, 'R')
        pdf.cell(45, 8, f"Rp {row['subtotal']:,.0f}", 1, 1, 'R')
    
    # Total
    pdf.ln(5)
    pdf.set_font("Arial", 'B', 12)
    pdf.cell(140, 10, "GRAND TOTAL", 0, 0, 'R')
    pdf.cell(45, 10, f"Rp {master_row['total']:,.0f}", 1, 1, 'R')
    
    return pdf.output(dest='S').encode('latin-1')


def show_transaction_history():
    t = st.session_state.theme
    st.markdown(f"<h4 style='color:{t['primary_color']}; font-family:Orbitron;'>📜 TRANSACTION HISTORY</h4>", unsafe_allow_html=True)
    
    #search_id = st.text_input("CARI ID INVOICE", placeholder="TRX-XXXX")

    conn = get_connection()
    df_p = pd.read_sql("SELECT id FROM trans_master", conn)
    search_id = st.selectbox("CARI ID INVOICE", df_p['id'], index=None, placeholder="TRX-XXXX")

    
    if search_id:
        conn = get_connection()
        master = pd.read_sql("SELECT * FROM trans_master WHERE id = ?", conn, params=(search_id,))
        
        if not master.empty:
            st.success(f"Invoice Ditemukan")
            detail = pd.read_sql("SELECT * FROM trans_detail WHERE trans_id = ?", conn, params=(search_id,))
            
            # Area Preview
            st.table(detail[['product_name', 'qty', 'price', 'subtotal']])
            
            # Tombol Cetak (Reprint)
            pdf_data = generate_invoice_pdf(master.iloc[0], detail)
            
            #pdf_bytes = bytes(pdf.output()) 
            
            st.download_button(
                label="📥 REPRINT INVOICE (PDF)",
                data=pdf_data,
                file_name=f"Invoice_{search_id}.pdf",
                mime="application/pdf"
            )

        else:
            st.error("Invoice tidak ditemukan.")


def show_product_search():
    t = st.session_state.theme
    st.markdown(f"<h2 style='color:{t['primary_color']}; font-family:Orbitron;'>🔍 PRODUCT FINDER</h2>", unsafe_allow_html=True)
    
    tab1, tab2 = st.tabs(["[ SCAN BARCODE ]", "[ MANUAL SEARCH ]"])
    
    conn = get_connection()
    
    with tab1:
        st.info("Simulasi: Masukkan angka barcode atau gunakan scanner fisik yang terhubung ke keyboard.")
        barcode_input = st.text_input("SCANNING AREA...", placeholder="Arahkan kursor di sini & scan...", key="barcode_scan")
        
        if barcode_input:
            # Cari berdasarkan barcode (Simulasi: ID produk digunakan sebagai barcode)
            res = pd.read_sql("SELECT * FROM products WHERE id = ?", conn, params=(barcode_input,))
            if not res.empty:
                display_product_card(res.iloc[0])
            else:
                st.warning("Produk tidak terdaftar!")

    with tab2:
        search_name = st.text_input("Cari Nama Produk...")
        if search_name:
            res = pd.read_sql("SELECT * FROM products WHERE name LIKE ?", conn, params=(f'%{search_name}%',))
            if not res.empty:
                for _, row in res.iterrows():
                    display_product_card(row)
            else:
                st.write("Produk tidak ditemukan.")

def display_product_card(row):
    t = st.session_state.theme
    # Card visual untuk produk yang ditemukan
    st.markdown(f"""
        <div style="border: 1px solid {t['primary_color']}; padding: 15px; border-radius: 10px; background: rgba(255,255,255,0.05); margin-bottom: 10px;">
            <h3 style="color:{t['primary_color']}; margin:0;">{row['name']}</h3>
            <p style="margin:5px 0;">Kategori: {row['category']} | <b>Stok: {row['stock']}</b></p>
            <h2 style="margin:0; color:white;">Rp {row['price']:,.0f}</h2>
        </div>
    """, unsafe_allow_html=True)



def generate_invoice_pdf(master_row, detail_df):
    pdf = FPDF()
    pdf.add_page()
    
    # --- Header Struk (FUTUREMART 2026) ---
    pdf.set_font("helvetica", 'B', 16)
    pdf.cell(0, 10, "FUTUREMART 2026", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    pdf.set_font("helvetica", '', 10)
    pdf.cell(0, 5, "Sistem Kasir Futuristik Terintegrasi", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    pdf.ln(10)
    
    # --- Informasi Transaksi (ID, Tanggal, Kasir) ---
    pdf.set_font("helvetica", 'B', 10)
    pdf.cell(0, 7, f"ID INVOICE : {master_row['id']}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font("helvetica", '', 10)
    pdf.cell(0, 7, f"Tanggal    : {master_row['date']}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.cell(0, 7, f"Kasir      : {master_row['cashier']}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(5)
    
    # --- Tabel Header (Judul Kolom) ---
    pdf.set_font("helvetica", 'B', 10)
    pdf.set_fill_color(200, 200, 200)
    
    # Tentukan lebar kolom agar rapi dan berjejer
    w_name = 85
    w_qty = 20
    w_price = 40
    w_sub = 45

    # Gunakan XPos.RIGHT agar kursor pindah ke kanan setelah setiap cell
    pdf.cell(w_name, 8, "Nama Produk", border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='L', fill=True)
    pdf.cell(w_qty, 8, "Qty", border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C', fill=True)
    pdf.cell(w_price, 8, "Harga", border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C', fill=True)
    
    # Gunakan XPos.LMARGIN untuk kolom terakhir agar pindah ke baris baru setelahnya
    pdf.cell(w_sub, 8, "Subtotal", border=1, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C', fill=True)
    
    # --- Tabel Detail (Data Produk) ---
    pdf.set_font("helvetica", '', 10)
    for _, row in detail_df.iterrows():
        # Data Nama Produk (Pindah kanan)
        pdf.cell(w_name, 8, str(row['product_name']), border=1, new_x=XPos.RIGHT, new_y=YPos.TOP)
        # Data Qty (Pindah kanan, Rata Tengah)
        pdf.cell(w_qty, 8, str(row['qty']), border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C')
        # Data Harga (Pindah kanan, Rata Kanan)
        pdf.cell(w_price, 8, f"Rp {row['price']:,.0f}", border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='R')
        # Data Subtotal (Pindah baris, Rata Kanan)
        pdf.cell(w_sub, 8, f"Rp {row['subtotal']:,.0f}", border=1, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='R')
        
    # --- Total ---
    pdf.ln(5)
    pdf.set_font("helvetica", 'B', 12)
    # Total Lebar untuk label "GRAND TOTAL"
    total_label_width = w_name + w_qty + w_price
    pdf.cell(total_label_width, 10, "GRAND TOTAL  ", new_x=XPos.RIGHT, new_y=YPos.TOP, align='R')
    pdf.cell(w_sub, 10, f"Rp {master_row['total']:,.0f}", border='B', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='R')
    
    # Mengembalikan dalam bentuk bytes untuk Streamlit (Menghindari error bytearray)
    return bytes(pdf.output())
   

def get_monthly_report():
    conn = get_connection()
    # Query untuk merekap data per bulan
    query = """
    SELECT 
        strftime('%Y-%m', date) as Bulan,
        COUNT(id) as Total_Transaksi,
        SUM(total + discount) as Pendapatan_Kotor,
        SUM(discount) as Total_Diskon,
        SUM(total) as Pendapatan_Bersih
    FROM trans_master
    GROUP BY Bulan
    ORDER BY Bulan DESC
    """
    df_report = pd.read_sql(query, conn)
    
    # Hitung Total Keseluruhan (Grand Total) untuk baris bawah
    if not df_report.empty:
        summary_row = pd.DataFrame({
            'Bulan': ['TOTAL'],
            'Total_Transaksi': [df_report['Total_Transaksi'].sum()],
            'Pendapatan_Kotor': [df_report['Pendapatan_Kotor'].sum()],
            'Total_Diskon': [df_report['Total_Diskon'].sum()],
            'Pendapatan_Bersih': [df_report['Pendapatan_Bersih'].sum()]
        })
        df_report = pd.concat([df_report, summary_row], ignore_index=True)
    
    return df_report

def to_excel(df):
    output = io.BytesIO()
    # Menulis ke Excel menggunakan engine openpyxl
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Laporan_Keuangan')
    return output.getvalue()

def play_ai_voice(text):
    # JavaScript untuk memicu suara asisten AI
    js_code = f"""
        <script>
        var msg = new SpeechSynthesisUtterance("{text}");
        msg.lang = 'id-ID';  // Bahasa Indonesia
        msg.pitch = 1.2;     // Sedikit tinggi agar terdengar futuristik
        msg.rate = 1.0;
        window.speechSynthesis.speak(msg);
        </script>
    """
    st.components.v1.html(js_code, height=0)

def play_tech_chime():
    # Suara notifikasi modern yang bersih
    audio_url = "https://actions.google.com/sounds/v1/alarms/beep_short.ogg" # Ganti dengan URL suara synth yang Anda suka
    audio_html = f"""
        <audio autoplay>
            <source src="{audio_url}" type="audio/ogg">
        </audio>
    """
    st.components.v1.html(audio_html, height=0)

def play_ai_voice_premium(text):
    js_code = f"""
        <script>
        function speakNow() {{
            const msg = new SpeechSynthesisUtterance("{text}");
            const voices = window.speechSynthesis.getVoices();
            
            // MENCARI SUARA GOOGLE INDONESIA (Reguler/Natural)
            // Biasanya bernama 'Google Bahasa Indonesia'
            let selectedVoice = voices.find(v => v.name === 'Google Bahasa Indonesia');
            
            // Fallback jika Google Voice tidak ditemukan (pakai ID umum)
            if (!selectedVoice) {{
                selectedVoice = voices.find(v => v.lang.includes('id-ID'));
            }}

            if (selectedVoice) {{
                msg.voice = selectedVoice;
            }}

            // PENGATURAN AGAR TIDAK KAKU
            msg.pitch = 1.0;  // Normal (tidak terlalu tinggi/rendah)
            msg.rate = 1.0;   // Kecepatan normal manusia bercakap
            msg.volume = 1.0;
            msg.lang = 'id-ID';

            window.speechSynthesis.speak(msg);
        }}

        if (window.speechSynthesis.getVoices().length === 0) {{
            window.speechSynthesis.onvoiceschanged = speakNow;
        }} else {{
            speakNow();
        }}
        </script>
    """
    st.components.v1.html(js_code, height=0)


def play_alert_sound():
    # Menggunakan suara 'beep' pendek dalam format base64 agar tidak perlu file eksternal
    audio_html = """
        <audio autoplay>
            <source src="https://www.soundjay.com/buttons/beep-01a.mp3" type="audio/mpeg">
        </audio>
    """
    st.components.v1.html(audio_html, height=0)


def check_low_stock_alerts(threshold=5):
    conn = get_connection()
    low_stock_df = pd.read_sql("SELECT name FROM products WHERE stock <= ?", conn, params=(threshold,))
    
    if not low_stock_df.empty:
        # Mengambil satu nama produk untuk disebut oleh AI
        #produk_list = ", ".join(low_stock_df['name'].tolist()[:6]) # Ambil [:6] produk pertama saja agar tidak kepanjangan
        produk_list = ", ".join(low_stock_df['name'])
        #pesan = f"Peringatan sistem. Stok {produk_list} hampir habis. Segera lakukan pengisian."
        pesan = f"Maaf, peringatan stok. Produk {produk_list} sudah mencapai batas minimum. Mohon segera melakukan pemesanan ulang."
        
        
        #play_ai_voice(pesan)
        #play_tech_chime()
        play_ai_voice_premium(pesan)
        
        # UI Alert yang cantik
        st.toast(f"🚨 {pesan}", icon="⚠️")
        
        # Tampilan Visual (Blinking Red)
        st.markdown(f'''
            <div style="background: rgba(255,0,0,0.2); border-left: 5px solid red; padding: 10px; border-radius: 5px;">
                <h4 style="color: #ff4b4b; margin:0;">🚨 SYSTEM NOTIFICATION</h4>
                <p style="margin:0;">{pesan}</p>
            </div>
        ''', unsafe_allow_html=True)



def show_financial_report():
    t = st.session_state.theme
    st.markdown(f"<h2 style='color:{t['primary_color']}; font-family:Orbitron;'>📊 FINANCIAL REPORT</h2>", unsafe_allow_html=True)
    
    report_df = get_monthly_report()
    
    if not report_df.empty:
        # Tampilan tabel di Streamlit
        st.dataframe(report_df.style.format({
            'Pendapatan_Kotor': 'Rp {:,.0f}',
            'Total_Diskon': 'Rp {:,.0f}',
            'Pendapatan_Bersih': 'Rp {:,.0f}'
        }), width='stretch')
        
        # Tombol Download Excel
        excel_data = to_excel(report_df)
        st.download_button(
            label="📈 DOWNLOAD LAPORAN EXCEL",
            data=excel_data,
            file_name=f"Laporan_FutureMart_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.info("Belum ada data transaksi untuk dilaporkan.")



# --- 5. SETTINGS (Termasuk Warna Body) ---
def show_settings():
    st.title("⚙️ SYSTEM CONFIG")
    t = st.session_state.theme
    st.markdown('<div class="glass-card">', unsafe_allow_html=True)
    
    st.subheader("Visual Identity")
    c1, c2, c3 = st.columns(3)
    new_primary = c1.color_picker("Top Bar", t['top_bar'])
    new_bg = c2.color_picker("Sidebar/Base", t['bg_color'])
    new_body = c3.color_picker("Main Page Body", t['body_color'])
    new_second = new_bg
    
    _="""
    if st.button("APPLY SETTINGS"):
        st.session_state.theme.update({
            'primary_color': new_primary, 
            'bg_color': new_bg, 
            'body_color': new_body,
            'top_bar': new_primary,
            'second_color': new_second
        })
        st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)
    """

    if st.button("SIMPAN PERUBAHAN TEMA"):
        st.session_state.theme['bg_color'] = new_bg
        st.session_state.theme['body_color'] = new_body
        st.session_state.theme['primary_color'] = new_primary
        st.session_state.theme['top_bar'] = new_primary
        st.session_state.theme['second_color'] = new_second
        
        # PERMANENKAN KE FILE
        save_theme(st.session_state.theme)
    
        st.success("Tema Berhasil Disimpan Permanen!")
        st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

def login_user(username, password):
    conn = get_connection()
    cursor = conn.cursor()
    
    # Cari user berdasarkan username
    cursor.execute("SELECT password, role FROM users WHERE username = ?", (username,))
    result = cursor.fetchone()
    conn.close()

    if result:
        db_password = result[0]
        db_role = result[1]
        # Bandingkan hash password input dengan hash di database
        if hash_password(password) == db_password:
            return True, db_role
    return False, None

def get_stock_predictions():
    conn = get_connection()
    # Ambil data penjualan 30 hari terakhir
    date_limit = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    
    
    query = """
        SELECT product_name, SUM(qty) as total_sold 
        FROM trans_detail 
        JOIN trans_master ON trans_detail.trans_id = trans_master.id
        WHERE trans_master.date >= ?
        GROUP BY product_name
    """
    df_sales = pd.read_sql(query, conn, params=[date_limit])
    df_products = pd.read_sql("SELECT name, stock FROM products", conn)
    conn.close()

    # Gabungkan data
    df_pred = pd.merge(df_products, df_sales, left_on='name', right_on='product_name', how='left').fillna(0)
    
    # Hitung Kecepatan Penjualan (Daily Velocity)
    df_pred['daily_velocity'] = df_pred['total_sold'] / 30
    
    # Hitung Sisa Hari (ETA Out of Stock)
    def calculate_eta(row):
        if row['daily_velocity'] <= 0: return 999 # Stok aman/tidak laku
        return math.floor(row['stock'] / row['daily_velocity'])

    df_pred['days_left'] = df_pred.apply(calculate_eta, axis=1)
    return df_pred

def show_accounting():
    st.markdown("<h1 style='color:var(--primary); font-family:Orbitron;'>FINANCIAL & AUDIT</h1>", unsafe_allow_html=True)
    conn = get_connection()
    df_m = pd.read_sql("SELECT * FROM trans_master", conn)
    
    # --- TAB LAPORAN ---
    tab1, tab2 = st.tabs(["📊 Jurnal Umum", "👥 Performa Kasir"])
    
    with tab1:
        st.markdown('<div class="glass-card">', unsafe_allow_html=True)
        st.subheader("Data Transaksi")
        st.dataframe(df_m.sort_values('date', ascending=False), width='stretch')
        st.markdown('</div>', unsafe_allow_html=True)

    with tab2:
        st.subheader("Analisis Produktivitas Karyawan")
        # Agregasi data per Kasir
        cashier_perf = df_m.groupby('cashier').agg({
            'id': 'count',
            'total': 'sum'
        }).rename(columns={'id': 'Total Transaksi', 'total': 'Total Omzet'}).reset_index()
        
        c1, c2 = st.columns([1, 1.5])
        with c1:
            st.markdown('<div class="glass-card">', unsafe_allow_html=True)
            st.write("Ranking Penjualan")
            st.table(cashier_perf.sort_values('Total Omzet', ascending=False))
            st.markdown('</div>', unsafe_allow_html=True)
        
        with c2:
            st.markdown('<div class="glass-card">', unsafe_allow_html=True)
            fig = px.bar(cashier_perf, x='cashier', y='Total Omzet', 
                         color='Total Transaksi', title="Perbandingan Omzet Kasir",
                         template="plotly_dark", color_continuous_scale='Viridis')
            fig.update_layout(paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)')
            st.plotly_chart(fig, width='stretch')
            st.markdown('</div>', unsafe_allow_html=True)

def show_forecasting():
    st.markdown("<h2 style='color:var(--primary); font-family:Orbitron;'>🔮 INVENTORY PREDICTION</h2>", unsafe_allow_html=True)

    check_low_stock_alerts(threshold=5)    
    
    df_pred = get_stock_predictions()
    
    # Filter Barang Kritikal (Habis < 7 hari)
    critical = df_pred[df_pred['days_left'] <= 7].sort_values('days_left')
    
    if not critical.empty:
        st.markdown('<div class="glass-card">', unsafe_allow_html=True)
        st.subheader("🚨 Restock Priority (ETA < 7 Days)")
        for _, row in critical.iterrows():
            st.markdown(f"""
                <div style="display:flex; justify-content:space-between; border-bottom:1px solid rgba(255,255,255,0.1); padding:10px 0;">
                    <span>{row['name']}</span>
                    <span class="critical-alert">Habis dalam {row['days_left']} hari</span>
                </div>
            """, unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

    # Visualisasi Proyeksi Stok
    st.markdown('<div class="glass-card">', unsafe_allow_html=True)
    st.subheader("Stock Endurance Analysis")
    fig = px.bar(df_pred, x='name', y='days_left', color='days_left',
                 labels={'days_left': 'Sisa Hari (Estimasi)', 'name': 'Produk'},
                 color_continuous_scale='RdYlGn', template="plotly_dark")
    fig.add_hline(y=7, line_dash="dash", line_color="red", annotation_text="Danger Zone")
    fig.update_layout(paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)')
    st.plotly_chart(fig, width='stretch')
    st.markdown('</div>', unsafe_allow_html=True)

# --- FUNGSI SUARA PREMIUM ---
def play_ai_voice_premiumx(text):
    # Menggunakan suara yang lebih jernih dengan kontrol pitch yang pas
    js_code = f"""
        <script>
        function speakNow() {{
            const msg = new SpeechSynthesisUtterance("{text}");
            
            // Mengambil semua daftar suara yang tersedia di sistem/browser
            const voices = window.speechSynthesis.getVoices();
            
            // Mencari suara wanita Indonesia dengan urutan prioritas:
            // 1. Google Bahasa Indonesia (Sangat jernih)
            // 2. Microsoft Gadis (Natural)
            // 3. Indonesian Standard
            const femaleVoice = voices.find(v => 
                v.lang.includes('id') && 
                (v.name.includes('Google') || v.name.includes('Gadis') || v.name.includes('Siti'))
            );

            if (femaleVoice) {{
                msg.voice = femaleVoice;
            }}

            // PENGATURAN KARAKTER SUARA
            msg.pitch = 1.3;  // Skala 0 - 2. Angka 1.3 memberikan kesan lebih feminin/lembut
            msg.rate = 0.95;  // Kecepatan sedikit dikurangi (normal = 1.0) agar artikulasi jelas
            msg.volume = 1.0; // Volume maksimal
            msg.lang = 'id-ID';

            window.speechSynthesis.speak(msg);
        }}

        // Mengatasi masalah daftar suara yang belum dimuat (Asynchronous)
        if (window.speechSynthesis.getVoices().length === 0) {{
            window.speechSynthesis.onvoiceschanged = speakNow;
        }} else {{
            speakNow();
        }}
        </script>
    """
    st.components.v1.html(js_code, height=0)



def get_next_id():
    with get_connection() as conn:
        res = conn.execute("SELECT MAX(id) FROM products").fetchone()[0]
    return (res + 1) if res else 1

def show_category_ui():
    # Menggunakan koneksi database
    with get_connection() as conn:
        df_c = pd.read_sql("SELECT id AS 'ID', name AS 'Nama Kategori' FROM categories", conn)
    
    st.subheader("Daftar Kategori")
    # Menampilkan tabel statis yang bersih
    st.dataframe(df_c, use_container_width=True, hide_index=True)

    c_add, c_edit = st.columns(2)
    
    # --- BAGIAN TAMBAH KATEGORI ---
    with c_add:
        with st.container(border=True):
            st.write("**Tambah Kategori**")
            new_cat = st.text_input("Nama Kategori Baru", key="input_new_cat")
            if st.button("Simpan Kategori", use_container_width=True):
                if new_cat:
                    try:
                        with get_connection() as conn:
                            conn.execute("INSERT INTO categories (name) VALUES (?)", (new_cat,))
                            conn.commit()
                        play_ai_voice_premium(f"Kategori {new_cat} berhasil ditambahkan.")
                        st.success("Tersimpan!")
                        time.sleep(0.5)
                        st.rerun()
                    except:
                        st.error("Nama kategori sudah ada!")
                else:
                    st.warning("Isi nama kategori!")

    # --- BAGIAN EDIT & HAPUS KATEGORI ---
    with c_edit:
        with st.container(border=True):
            st.write("**Modifikasi Kategori**")
            cat_options = df_c['Nama Kategori'].tolist()
            target_cat = st.selectbox("Pilih Kategori", cat_options, index=None, placeholder="Pilih...", key="select_edit_cat")
            
            if target_cat:
                rename_cat = st.text_input("Ganti Nama", value=target_cat, key=f"rename_{target_cat}")
                
                col_u, col_d = st.columns(2)
                
                # Logika Update Nama Kategori & Produk Terkait
                if col_u.button("Update", use_container_width=True):
                    with get_connection() as conn:
                        # Update tabel kategori
                        conn.execute("UPDATE categories SET name=? WHERE name=?", (rename_cat, target_cat))
                        # Update semua produk yang menggunakan kategori lama ini (Relasi Sinkron)
                        conn.execute("UPDATE products SET category=? WHERE category=?", (rename_cat, target_cat))
                        conn.commit()
                    play_ai_voice_premium("Kategori berhasil diperbarui.")
                    st.rerun()
                
                # Logika Hapus dengan Proteksi Relasi
                if col_d.button("Hapus", use_container_width=True, type="primary"):
                    with get_connection() as conn:
                        # Cek apakah ada produk yang masih memakai kategori ini
                        count = conn.execute("SELECT COUNT(*) FROM products WHERE category=?", (target_cat,)).fetchone()[0]
                        
                        if count > 0:
                            st.error(f"Ditolak! {count} produk masih menggunakan kategori ini.")
                            play_ai_voice_premium("Kategori tidak bisa dihapus karena masih digunakan oleh produk.")
                        else:
                            conn.execute("DELETE FROM categories WHERE name=?", (target_cat,))
                            conn.commit()
                            play_ai_voice_premium(f"Kategori {target_cat} dihapus.")
                            st.rerun()

def show_inventory_system():
    t = st.session_state.theme
    st.markdown(f"<h2 style='color:{t['primary_color']}; font-family:Orbitron;'>🏬 INVENTORY CONTROL</h2>", unsafe_allow_html=True)
    
    # 1. Inisialisasi Session State agar form tidak tertutup saat klik tombol
    if 'editing_product' not in st.session_state:
        st.session_state.editing_product = None
    if 'last_action' not in st.session_state:
        st.session_state.last_action = None

    tab_prod, tab_cat = st.tabs(["📦 MANAJEMEN PRODUK", "🗂️ MANAJEMEN KATEGORI"])

    # ================= TAB 1: MANAJEMEN PRODUK =================
    with tab_prod:
        # Load kategori dari DB
        with get_connection() as conn:
            cat_list = pd.read_sql("SELECT name FROM categories", conn)['name'].tolist()
        
        # --- SUB-TAB: TAMBAH PRODUK ---
        with st.expander("➕ TAMBAH PRODUK BARU", expanded=False):
            next_id = get_next_id()
            with st.form("add_product_form", clear_on_submit=True):
                col1, col2 = st.columns(2)
                p_name = col1.text_input("Nama Produk")
                p_cat = col1.selectbox("Kategori", options=cat_list if cat_list else ["Lainnya"])
                p_cost = col2.number_input("Harga Modal (Rp)", min_value=0)
                p_price = col2.number_input("Harga Jual (Rp)", min_value=0)
                p_stock = col2.number_input("Stok Awal", min_value=0)
                p_barcode = col1.text_input("Barcode")
                
                if st.form_submit_button("SIMPAN PRODUK", use_container_width=True):
                    if p_price >= p_cost and p_name:
                        with get_connection() as conn:
                            conn.execute("INSERT INTO products VALUES (?,?,?,?,?,?,?)", 
                                         (next_id, p_name, p_cat, p_price,p_stock, p_barcode,p_cost))
                            conn.commit()
                        play_ai_voice_premium(f"Produk {p_name} berhasil disimpan.")
                        st.success("Tersimpan!")
                        st.rerun()
                    else:
                        st.error("Periksa nama dan margin harga!")

        # --- SUB-TAB: DAFTAR & EDIT ---
        st.subheader("Daftar & Edit Produk")
        with get_connection() as conn:
            df_p = pd.read_sql("SELECT * FROM products", conn)
        
        # Menggunakan session state untuk selectbox agar tetap terpilih
        selected_name = st.selectbox(
            "Pilih Produk", 
            df_p['name'].tolist(), 
            index=None, 
            placeholder="Cari produk..."
        )

        if selected_name:
            # Simpan data ke session state agar stabil
            st.session_state.editing_product = df_p[df_p['name'] == selected_name].iloc[0].to_dict()
            p = st.session_state.editing_product

            with st.container(border=True):
                c_a, c_b = st.columns(2)
                # Gunakan key unik dari ID agar tidak tertukar di memory
                u_name = c_a.text_input("Nama", value=p['name'], key=f"edit_nm_{p['id']}")
                u_cat = c_a.selectbox("Kategori", cat_list, index=cat_list.index(p['category']) if p['category'] in cat_list else 0, key=f"edit_ct_{p['id']}")
                u_cost = c_b.number_input("Modal", value=float(p['cost_price']), key=f"edit_cs_{p['id']}")
                u_price = c_b.number_input("Jual", value=float(p['price']), key=f"edit_pr_{p['id']}")
                u_stock = c_b.number_input("Stok", value=int(p['stock']), key=f"edit_st_{p['id']}")
                
                col_btn1, col_btn2 = st.columns(2)
                
                if col_btn1.button("💾 UPDATE", use_container_width=True):
                    if u_price >= u_cost:
                        with get_connection() as conn:
                            conn.execute("""UPDATE products SET name=?, category=?, price=?, cost_price=?, stock=? 
                                         WHERE id=?""", (u_name, u_cat, u_price, u_cost, u_stock, p['id']))
                            conn.commit()
                        play_ai_voice_premium(f"Data {u_name} berhasil diperbarui. Check Baby check wan thru trhee")
                        st.success("Update Berhasil!")
                        time.sleep(0.5)
                        st.rerun()

                if col_btn2.button("🗑️ DELETE", use_container_width=True, type="primary"):
                    with get_connection() as conn:
                        # Cek relasi child table (trans_detail)
                        rel = conn.execute("SELECT COUNT(*) FROM trans_detail WHERE product_id=?", (p['id'],)).fetchone()[0]
                        if rel > 0:
                            st.error(f"Gagal! Ada {rel} transaksi terkait.")
                            play_ai_voice_premium("Penghapusan ditolak karena relasi transaksi.")
                        else:
                            conn.execute("DELETE FROM products WHERE id=?", (p['id'],))
                            conn.commit()
                            st.success("Terhapus!")
                            time.sleep(0.5)
                            st.rerun()

    # ================= TAB 2: MANAJEMEN KATEGORI =================
    with tab_cat:
        # (Logika kategori tetap sama, namun gunakan st.rerun() setelah commit)
        show_category_ui() # Fungsi kategori dipisah agar rapi

# (Main Routing Logic tetap sama seperti sebelumnya)
def app_supermarket():
    st.set_page_config(page_title="FutureMart 2026", layout="wide")

    
    # 1. Pastikan variabel state tersedia
    if 'logged_in' not in st.session_state: st.session_state.logged_in = False
    if 'cart' not in st.session_state: st.session_state.cart = []
    
# Inisialisasi tema dari file atau default
    if 'theme' not in st.session_state:
        saved_theme = load_theme()
        if saved_theme:
            st.session_state.theme = saved_theme
        else:
            st.session_state.theme = {
                'bg_color': '#adb99e', 
                'second_color': '#adb99e', 
                'top_bar': '#7d7f72', 
                'body_color': '#bbc9a7', 
                'primary_color': '#7d7f72', 
                'font_family': 'Rajdhani', 
                'font_size': '16px'            
            }

    # 3. Pindahkan apply_theme() ke LUAR blok 'if' agar selalu dijalankan
    apply_theme()

# Inisialisasi tema dari file atau default
    if 'theme' not in st.session_state:
        saved_theme = load_theme()
               
        if saved_theme:
            st.session_state.theme = saved_theme
        else:
            st.session_state.theme = {
                'bg_color': '#adb99e', 
                'second_color': '#adb99e', 
                'top_bar': '#7d7f72', 
                'body_color': '#bbc9a7', 
                'primary_color': '#7d7f72', 
                'font_family': 'Rajdhani', 
                'font_size': '16px'
            }

      
    if not st.session_state.logged_in:
   
            
            st.components.v1.html(
                """
                <script>
                    var input = window.parent.document.querySelectorAll("input[type=text]");
                    for (var i = 0; i < input.length; ++i) {
                        if (input[i].getAttribute('aria-label') === "Username") {
                            input[i].focus();
                        }
                    }
                </script>
                """,
                height=0,
            )    

   
            st.markdown("<h1 style='text-align:center;'>LOGIN SYSTEM</h1>", unsafe_allow_html=True)
            _, col, _ = st.columns([1,1,1])
            with col.form("login"):
                u = st.text_input("Username", placeholder="Enter username...")
                p = st.text_input("Password", type="password", placeholder="Enter password...")
                if st.form_submit_button("MASUK"):
                    is_valid, role = login_user(u, p)
                    if is_valid:
                        st.session_state.logged_in = True
                        st.session_state.username = u
                        st.session_state.user_role = role         
                        add_log(u, "LOGIN") # <--- CATAT LOGS LOGIN USER                 
                        st.success(f"Selamat Datang, {u} ({role})")
                        st.rerun()
                    else:
                        st.error("Username atau Password Salah!")



    else:          
        with st.sidebar:


            st.markdown("""
                <style>
                :root {
                    /* Tetap sinkronkan variabel jika diperlukan */
                    --secondary-background-color: transparent !important;
                }

                /* Membuat latar belakang widget menu/input menjadi transparan */
                div[data-baseweb="input"], 
                div[data-baseweb="select"], 
                div[data-baseweb="popover"], /* Menargetkan menu popover/dropdown */
                .stSecondaryButton,
                div[data-testid="stSidebar"] > div:first-child {
                    background-color: transparent !important;
                    background: transparent !important;
                }

                /* Opsional: Menghilangkan border agar benar-benar terlihat 'bersih' */
                div[data-baseweb="input"] {
                    border: 1px solid rgba(255, 255, 255, 0.2) !important;
                }
                </style>
                """, unsafe_allow_html=True)  
            #st.markdown(f"<h1 style='color:var(--primary); font-family:Orbitron;'>FM-POS</h1>", unsafe_allow_html=True)
            st.write(f"User: **{st.session_state.username}** ({st.session_state.user_role})")


            selected = sac.menu([
                sac.MenuItem('home', icon='house-fill'),
                sac.MenuItem('Dashboard', icon='bar-chart-fill'),
                sac.MenuItem('Operation', icon='box-fill', children=[
                    sac.MenuItem('Transaction', icon='bank2'),
                    sac.MenuItem('Log transaction', icon='git', children=[
                        sac.MenuItem('Riwayat', icon='hourglass-split'),   
                        sac.MenuItem('Cari Produk', icon='eye-fill'),                      
                    ]),
                    ]),
                sac.MenuItem('Forecasting', icon='sliders'),
                sac.MenuItem('Inventory system', icon='basket'),
                sac.MenuItem('Laporan Keuangan', icon='journal-album'),
                sac.MenuItem('Accounting', icon='book'),
                sac.MenuItem('Activity Logs', icon='substack'),
                sac.MenuItem('Administrator', icon='person-fill-gear', children=[
                    sac.MenuItem('User Mgmt', icon='person-circle'),
                    sac.MenuItem('Backup_DB_Online', icon='database-fill'),
                    sac.MenuItem('Settings', icon='gear'),
                    ])
                ], variant='left-bar', color='white',)


            if st.button("LOGOUT"):
                add_log(st.session_state.username, "LOGOUT") # <--- CATAT LOGOUT
                st.session_state.logged_in = False
                for key in list(st.session_state.keys()):
                    del st.session_state[key]                
                st.rerun()
      

        # Routing Logic
        match selected:


            case "Dashboard": show_dashboard()
                #st.write("Dashboard (Admin Only)")
            case "Forecasting": show_forecasting()
                #st.write("Forecasting (Admin Only)")
            case "Transaction": show_pos()
            case "Inventory system": show_inventory_system()
            case "Riwayat": show_transaction_history()
            case "Cari Produk": show_product_search()
            case "Laporan Keuangan": show_financial_report()
            
                #st.write("Backup_DB_Online (Admin Only)")
            case "Accounting": show_accounting() 
            case "Activity Logs": show_activity_logs()    

            case "User Mgmt":
                if st.session_state.user_role == 'Admin':
                    show_user_mgmt()
                else:
                     st.write("User Mgmt (Admin Only)")   

            case "Settings": 
                if st.session_state.user_role == 'Admin':
                    show_settings()
                else:
                     st.write("Settings (Admin Only)")   
            case "Backup_DB_Online":
                if st.session_state.user_role == 'Admin':
                    show_database_tools()
                else:
                     st.write("Backup_DB_Online (Admin Only)")   
            
            case _: st.write("Pilih menu untuk memulai.")



if __name__ == "__main__":

    st.markdown("""
        <style>
        :root {
            /* Tetap sinkronkan variabel jika diperlukan */
            --secondary-background-color: transparent !important;
        }

        /* Membuat latar belakang widget menu/input menjadi transparan */
        div[data-baseweb="input"], 
        div[data-baseweb="select"], 
        div[data-baseweb="popover"], /* Menargetkan menu popover/dropdown */
        .stSecondaryButton,
        div[data-testid="stSidebar"] > div:first-child {
            background-color: transparent !important;
            background: transparent !important;
        }

        /* Opsional: Menghilangkan border agar benar-benar terlihat 'bersih' */
        div[data-baseweb="input"] {
            border: 1px solid rgba(255, 255, 255, 0.2) !important;
        }
        </style>
        """, unsafe_allow_html=True)  
    app_supermarket()
